import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NFL"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/{excelFile}.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NFLScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(NFLScraper, self).__init__(executable_path="./geckodriver")
        self.games = []
        self.playersStats = []

    def openFoxBet(self) -> None:
        self.get("https://mtairycasino.foxbet.com/#/american_football/competitions")

    def getAllGames(self) -> None:
        games = WebDriverWait(self, 15).until(ec.presence_of_all_elements_located((By.CLASS_NAME,'event-schedule-additional-markets')))
        time.sleep(3.5)
        games = self.find_elements(By.CLASS_NAME,'event-schedule-additional-markets')
        for game in games:
            self.games.append(game.find_element(By.TAG_NAME, 'a').get_attribute('href'))

        print(f"Found: {len(self.games)} games")

    def openPlayerPropsTab(self) -> None:
        time.sleep(2)
        navbar = WebDriverWait(self, wait_time).until(ec.presence_of_element_located((By.CSS_SELECTOR, '.nav.nav-pills.market-groups.horizontalMenu__scroller')))
        tabs = navbar.find_elements(By.XPATH, './li')

        for tab in tabs:
            self.execute_script("arguments[0].scrollIntoView(true);", tab)
            try:
                if "player props" in tab.text.lower().strip():
                    self.execute_script("arguments[0].click();", tab)
                    self.execute_script("arguments[0].click();", tab.find_element(By.TAG_NAME, 'a'))
                    break
            except:
                print("No Props")

    def scrapGames(self, stats, startGame = 1, endGame ="all") -> None:
        self.startGame = startGame
        self.endGame = len(self.games) if endGame == "all" else endGame

        print(f"Scraping form Game: {startGame} until {endGame}")

        for game in self.games[self.startGame-1: self.endGame]:
            print("---------------------------------------")
            print(f"Scraping Game {self.games.index(game) + 1}")
            self.get(game)
            time.sleep(2)
            self.openPlayerPropsTab()
            time.sleep(2)
            self.loadStats(stats)

    def scrapPlayersInStat(self, statDiv, stat):

        time.sleep(1)

        sectionBody = WebDriverWait(statDiv, 15).until(ec.presence_of_element_located((By.CSS_SELECTOR, ".selectionBody.collapseToggle__content")))
        self.execute_script("arguments[0].scrollIntoView(true);", sectionBody)

        sectionBody = WebDriverWait(statDiv, 15).until(
            ec.presence_of_element_located((By.CSS_SELECTOR, ".selectionBody.collapseToggle__content")))


        overUnder = sectionBody.find_elements(By.CSS_SELECTOR, ".price")
        p = {
            "Player Name": stat.split(' - ')[0],
            "Stat": stat.split(' - ')[1],
            "Total": overUnder[0].find_element(By.CSS_SELECTOR, ".button__bet__title").text.split(" ")[1],
            "Over Odds": overUnder[0].find_element(By.CSS_SELECTOR, ".button__bet__odds").text,
            "Under Odds": overUnder[1].find_element(By.CSS_SELECTOR, ".button__bet__odds").text
        }
        self.playersStats.append(p)
        print(p)



    def loadStats(self, stats) -> None:
        allStatsDiv = WebDriverWait(self, wait_time).until(ec.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.markets-selectionView')))


        for statDiv in allStatsDiv:
            if any(stat in statDiv.find_element(By.CSS_SELECTOR, "span.groupHeader__titleText").text.strip() for stat in stats):
                print(statDiv.find_element(By.CSS_SELECTOR, "span.groupHeader__titleText").text.strip())
                statDiv = statDiv.find_element(By.CSS_SELECTOR, '.open.groupHeader.groupHeader--marketHeader')
                self.execute_script("arguments[0].scrollIntoView(true);", statDiv)
                self.execute_script("arguments[0].click();", statDiv)
                time.sleep(1)
                self.scrapPlayersInStat(statDiv.find_element(By.XPATH, './parent::div/parent::div/parent::div'), statDiv.find_element(By.CSS_SELECTOR, "span.groupHeader__titleText").text.strip())

with NFLScraper() as nfl:
    print("Opening FoxBet")
    nfl.openFoxBet()
    print("Getting All Games")
    nfl.getAllGames()

    nfl.scrapGames(["Passing Yds", "Passing TDs", "Rushing Yds", "Receiving Yds", "Receptions"],
        1,"all"
    )

    print("Extracting to Excel :) ")
    extractToExcel(nfl.playersStats, "nfl - data", "FoxBet")
    print("Done, closing..")