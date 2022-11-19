import openpyxl


from multiprocessing import freeze_support
from openpyxl import load_workbook

from undetected_chromedriver import Chrome
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

from selenium.webdriver.common.action_chains import ActionChains

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NBA"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/nba - data.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class BBALLScraper(Chrome):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(BBALLScraper, self).__init__(use_subprocess=True)
        self.playersStats = []
        self.actions = ActionChains(self)
        self.gameLinks = []

    def openPointsBet(self) -> None:
        self.get("https://pa.pointsbet.com/sports/basketball/NBA")
        WebDriverWait(self, 100).until(ec.presence_of_element_located(
            (By.CSS_SELECTOR, "div[identifier='sports_default_competition-main']"))
        )

    def wait_loading_finish(self) -> None:
        WebDriverWait(self,100).until(ec.invisibility_of_element_located((By.CSS_SELECTOR, 'span.fiajpul.f1j8lj9j.f18im1ze')))

    def getAllGamesLinks(self) -> None:
        """
            Get all game links
        """
        gameEvents = WebDriverWait(self, 30).until(
            ec.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-test="event"]'))
        )

        for game in gameEvents:
            try:
                self.gameLinks.append(
                    game.find_element(By.TAG_NAME, 'a').get_attribute('href')
                )
            except:
                pass

    def scrapProps(self, stats, propAccordion):
        time.sleep(1)

        print(propAccordion.text)


    def scrapGame(self, stats, link):
        self.get(link)

        accordions = WebDriverWait(self, 45).until(ec.presence_of_all_elements_located(
            (By.CSS_SELECTOR, 'button[name="accordionButton"]'))
        )

        props_accordion = False
        props_accordion_div = None
        for accordion in accordions:
            if accordion.text.lower().strip() == 'player props':
                accordion.click()
                props_accordion_div = accordion.find_element(By.XPATH, './..').find_element(By.TAG_NAME, 'div')
                break

        if props_accordion:
            self.scrapProps(stats, props_accordion_div)

    def scrapGames(self, stats , startGame = 1, endGame= "all") -> None:
        self.startGame = startGame - 1
        self.endGame = endGame
        if endGame == "all":
            self.endGame = len(self.gameLinks)

        for i in range(self.startGame, self.endGame):
            print(f"Scraping Game {i+1}")
            self.scrapGame(stats, self.gameLinks[i])

if __name__ == '__main__':
    freeze_support()

    with BBALLScraper() as BBALL:

        print("Opening Points Bet")
        BBALL.openPointsBet()
        print("Points Bet Opened Successfully..")

        print("Games are still being loaded")
        BBALL.wait_loading_finish()
        print("Games are loaded successfully")

        print("Getting All Games")
        BBALL.getAllGamesLinks()
        print(f"Found {len(BBALL.gameLinks)} games.")

        BBALL.scrapGames(["points scored",
             "assists",
             "rebounds",
             "pts scored + rebs + asts",
             "points scored + assists",
             "points scored + rebounds over/under",
             "assists + rebounds"], 1, "all")

        print("Extracting to Excel :) ")
        extractToExcel(BBALL.playersStats, "nba - data", "Barstool Raw")
        print("Done, closing..")
