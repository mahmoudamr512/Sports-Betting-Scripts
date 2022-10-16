import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NFL"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/{excelFile}.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NFLScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        options = webdriver.FirefoxOptions()
        options.add_argument('--disable-blink-features=AutomationControlled')
        self.BASE_URL = "https://sportsbook.caesars.com/us/pa/bet/americanfootball/events/all"
        super(NFLScraper, self).__init__(executable_path="./geckodriver", options=options)
        self.playersStats = []
        self.gameLinks = []

    def openCeasers(self) -> None:
        self.get(self.BASE_URL)

    def scroll_down(self) -> None:
        """A method for scrolling the page."""


        for i in range(13):
            # Scroll down to the bottom.
            self.execute_script("window.scrollBy(0, 750);", "")

            # Wait to load the page.
            time.sleep(1)


    def getAllGamesLinks(self) -> None:
        events = WebDriverWait(self, wait_time).until(ec.presence_of_element_located((By.CSS_SELECTOR, ".eventList")))
        events = WebDriverWait(self, wait_time).until(ec.presence_of_all_elements_located((
            By.CSS_SELECTOR, ".EventCard"
        )))
        time.sleep(4)

        self.scroll_down()

        time.sleep(1)
        events = WebDriverWait(self, wait_time).until(ec.presence_of_element_located((By.CSS_SELECTOR, ".eventList")))
        events = WebDriverWait(self, wait_time).until(ec.presence_of_all_elements_located((
            By.CSS_SELECTOR, ".EventCard"
        )))

        for event in events:
            self.gameLinks.append(f"{event.find_element(By.TAG_NAME, 'a').get_attribute('href')}")

    def openPropsTab(self, stats) -> None:
        time.sleep(2)

        navBar = WebDriverWait(self, 30).until(ec.presence_of_element_located((By.CSS_SELECTOR, 'ul.PillSlider')))
        self.execute_script("arguments[0].scrollIntoView(true);", navBar)

        for tab in navBar.find_elements(By.TAG_NAME, 'li'):
            if tab.text in ["Player Props"]:
                self.execute_script("arguments[0].scrollIntoView(true);", tab)
                self.execute_script("arguments[0].click();", tab)
                self.execute_script("arguments[0].click();", tab.find_element(By.TAG_NAME, 'button'))

                time.sleep(2)

                self.scrapStats(stats)

    def scrapGames(self, stats, startGame = 1, endGame = 'all') -> None:

        self.endGame = endGame
        self.startGame = startGame

        if endGame == 'all':
            self.endGame = len(self.gameLinks)

        counter = 1

        for link in self.gameLinks[(self.startGame)-1:self.endGame]:
            self.get(link)
            print(f"Scraping Game {counter}")
            self.openPropsTab(stats)
            print("Done")
            counter += 1

    def scrapStats(self, stats) -> None:
        mainDivs = WebDriverWait(self, wait_time).until(ec.presence_of_all_elements_located((By.CSS_SELECTOR, '.MultipleMarketsWithTabs')))

        for i in range(len(mainDivs)):
            try:
                div = mainDivs[i]
                headerSpan = div.find_element(By.CSS_SELECTOR, "span.title")
                if headerSpan.text.strip() in stats:
                    print(f"-----------------Scraping {headerSpan.text}-----------------------")
                    self.execute_script("arguments[0].scrollIntoView(true);", div)
                    time.sleep(0.25)
                    arrow =  div.find_element(By.CSS_SELECTOR, 'span.expanderHeaderRight')
                    self.execute_script("arguments[0].click();", arrow)
                    time.sleep(0.25)

                    teamsFlag = False

                    try:
                        div.find_element(By.CSS_SELECTOR, "ul.react-tabs__tab-list")
                        teamsFlag = True
                    except:
                        pass

                    if teamsFlag:
                        for team in mainDivs[i].find_element(By.CSS_SELECTOR, "ul.react-tabs__tab-list").find_elements(By.XPATH, './li'):
                            self.execute_script("arguments[0].click();", team)
                            time.sleep(0.5)

                            mainDivs = WebDriverWait(self, wait_time).until(
                                ec.presence_of_all_elements_located((By.CSS_SELECTOR, '.MultipleMarketsWithTabs')))
                            div = mainDivs[i]
                            playerGrid = div.find_element(By.CSS_SELECTOR, ".outcomesGrid.isPlayerProps")

                            dataArr = playerGrid.find_elements(By.XPATH, './div')

                            for j in range(0, len(dataArr), 3):
                                player = {
                                    "Player Name": dataArr[j].text,
                                    "Stat": headerSpan.text.strip(),
                                    "Total": dataArr[j + 1].text.split("\n")[0].split(" ")[1],
                                    "Over Odds": dataArr[j + 1].text.split("\n")[1],
                                    "Under Odds": dataArr[j + 2].text.split("\n")[1]
                                }

                                print(player)
                                self.playersStats.append(player)
                    else:
                        playerGrid = div.find_element(By.CSS_SELECTOR, ".outcomesGrid.isPlayerProps")

                        dataArr = playerGrid.find_elements(By.XPATH, './div')

                        for j in range(0, len(dataArr), 3):
                            player = {
                                "Player Name": dataArr[j].text,
                                "Stat": headerSpan.text.strip(),
                                "Total": dataArr[j+1].text.split("\n")[0].split(" ")[1],
                                "Over Odds": dataArr[j+1].text.split("\n")[1],
                                "Under Odds": dataArr[j+2].text.split("\n")[1]
                            }
                            self.playersStats.append(player)
                            print(player)

            except Exception as ex:
                pass


with NFLScraper() as nfl:
    print("Opening Caesars")
    nfl.openCeasers()
    print("Getting All Links")
    nfl.getAllGamesLinks()
    print(f"Found {len(nfl.gameLinks)} games.")
    nfl.scrapGames(["Total Passing Yards", "Total Rushing Yards", "Total Rushing Attempts",
                        "Total Receptions", "Total Receiving Yards"])

    print("Extracting to Excel :) ")
    extractToExcel(nfl.playersStats, "nfl - data", "Bet MGM")
    print("Done, closing..")
