import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NBA"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/nba - data.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    book = load_workbook(path)
    writer.book = book

    try:
        del book[excelTab]
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NBAScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(NBAScraper, self).__init__(executable_path="./geckodriver")
        self.playersStats = []

    def openBetOnline(self) -> None:
        self.get("https://bv2.digitalsportstech.com/betbuilder?sb=betonline")

    def loadNBATab(self) -> None:
        sports_div = WebDriverWait(self, wait_time).until(
            ec.visibility_of_element_located((By.CSS_SELECTOR, ".ligues-slider__list.sports")))
        sports = sports_div.find_elements(By.XPATH, './div')

        for sport in sports:
            if target_sport in sport.text:
                sport.click()
                break

    def scrapPlayersInStat(self, statDiv, stat):

        games = statDiv.find_elements(By.CSS_SELECTOR, 'div.main-stat__content')

        if self.endGame == 'all':
            self.endGame = len(games)

        for game in games[self.startGame-1:self.endGame]:
            time.sleep(1)
            print(f"Scrapping players of game {game.text}")
            self.execute_script("arguments[0].scrollIntoView(true);", game)
            game.click()
            playersList = WebDriverWait(game, 20).until(
                ec.visibility_of_element_located((By.TAG_NAME, 'app-main-stats-ou'))
            )

            for div in playersList.find_elements(By.CSS_SELECTOR, 'div.over-under-block__item'):
                self.playersStats.append(
                    {
                        "Player Name": div.find_element(By.CLASS_NAME, "over-under-block__player-name").text,
                        "Stat": stat.title(),
                        "Total": div.find_element(By.CSS_SELECTOR, 'span.highlight-text-color').text,
                        "Over Odds": div.find_elements(By.CSS_SELECTOR, ".over-under-block__selector-value")[0].text,
                        "Under Odds": div.find_elements(By.CSS_SELECTOR, ".over-under-block__selector-value")[1].text
                    }
                )

    def scrapStats(self, statDiv, stat):
        print(f"----------------------Scrapping all games in {stat} stat--------------------------")

        time.sleep(2)

        if 'main-stat--open' not in statDiv.find_element(By.XPATH, './div').get_attribute('outerHTML'):
            statDiv.find_element(By.CSS_SELECTOR, '.main-stats__item.main-stat').click()
            time.sleep(3)
        self.scrapPlayersInStat(statDiv, stat)

        time.sleep(1)


    def loadStats(self, stats, startGame = 1, endGame ='all') -> None:
        time.sleep(1)
        self.startGame = startGame
        self.endGame =  endGame
        mainStatsDiv = WebDriverWait(self, wait_time).until(ec.presence_of_element_located(
            (By.CSS_SELECTOR, "div.main-stats"))
        )
        allStatsDiv = mainStatsDiv.find_elements(By.TAG_NAME, 'app-main-stats-grouped')

        for stat in stats:
            for i in range(len(allStatsDiv)):
                if stat.lower() in allStatsDiv[i].text.lower().strip():
                    self.scrapStats(allStatsDiv[i], stat)
                    time.sleep(0.025)
                    allStatsDiv = mainStatsDiv.find_elements(By.TAG_NAME, 'app-main-stats-grouped')


with NBAScraper() as nba:
    print("Opening Betonline")
    nba.openBetOnline()
    print("Loading NBA Tab")
    nba.loadNBATab()
    nba.loadStats(
    ["Points", "Rebounds","Assists", "Pts + Reb + Ast"],
        1,1
    )

    print("Extracting to Excel :) ")
    extractToExcel(nba.playersStats, "nba - data", "BetOnline Raw")
    print("Done, closing..")