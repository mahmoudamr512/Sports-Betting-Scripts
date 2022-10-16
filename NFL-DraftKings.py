import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NFL"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NFLScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(NFLScraper, self).__init__(executable_path="./geckodriver")
        self.playersStats = []
        self.gameLinks = []

    def openDraftKings(self) -> None:
        self.BASE_URL = "https://sportsbook.draftkings.com/"
        self.get("https://sportsbook.draftkings.com/leagues/football/nfl")

    def getAllGames(self) -> None:
        table = WebDriverWait(self, 15).until(ec.presence_of_element_located((By.CSS_SELECTOR, ".sportsbook-table")))

        eventsRows =  table.find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, 'tr')

        i = 0

        while i < len(eventsRows):
            event = eventsRows[i]
            self.gameLinks.append(f"{event.find_element(By.TAG_NAME, 'a').get_attribute('href')}")
            i += 2

        print(f"Found {len(self.gameLinks)} games.")


    def scrapPlayers(self, headerDiv, stat):
        print(f"Scraping -----{stat}----")
        for tr in headerDiv.find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr"):
            overUnder = tr.find_elements(By.CSS_SELECTOR, ".sportsbook-outcome-cell__elements")
            self.playersStats.append({
                "Player Name": tr.find_element(By.CLASS_NAME, "sportsbook-row-name").text,
                "Stat": stat,
                "Total": tr.find_element(By.CSS_SELECTOR, ".sportsbook-outcome-cell__line").text,
                "Over Odds": overUnder[0].text,
                "Under Odds": overUnder[1].text
            })

    def scrapGame(self, stats) -> None:
        bottomSection = WebDriverWait(self, 15).until(
            ec.presence_of_element_located((By.CSS_SELECTOR, '.event-page-offers')))
        navbar = bottomSection.find_element(By.CSS_SELECTOR,
                                            ".sportsbook-tabbed-subheader")

        time.sleep(1)
        navbarLinks = navbar.find_elements(By.TAG_NAME, 'a')
        for i in range(1, len(navbarLinks)):
            self.execute_script(
                "arguments[0].scrollIntoView(true);", navbarLinks[i]
            )
            navbarLinks[i].click()
            time.sleep(1)

            container = self.find_element(By.CSS_SELECTOR, ".sportsbook-responsive-card-container__card.selected")

            for header in container.find_elements(By.CSS_SELECTOR, ".sportsbook-event-accordion__title.active"):
                if header.text.strip() in stats:
                    headerDiv = header.find_element(By.XPATH, "./parent::div/parent::div    ")
                    self.scrapPlayers(headerDiv, header.text.strip())
            navbar = bottomSection.find_element(By.CSS_SELECTOR,
                                                ".sportsbook-tabbed-subheader")
            navbarLinks = navbar.find_elements(By.TAG_NAME, 'a')

    def scrapGames(self, stats, startGame=1, endGame='all') -> None:

        self.endGame = endGame
        self.startGame = startGame

        if endGame == 'all':
            self.endGame = len(self.gameLinks)

        counter = 1

        for link in self.gameLinks[(self.startGame) - 1:self.endGame]:
            self.get(link)
            print(f"Scraping Game {counter}")
            self.scrapGame(stats)
            print("Done")
            counter += 1

with NFLScraper() as nfl:
    print("Opening DraftKings")
    nfl.openDraftKings()
    print("Getting All Links")
    nfl.getAllGames()
    nfl.scrapGames(["Pass Yds", "Rush Yds", "Rec Yds", "Receptions", "Rush Attempts"], 3, 4)

    print("Extracting to Excel :) ")
    extractToExcel(nfl.playersStats, "nfl - data", "DraftKings")
    print("Done, closing..")

