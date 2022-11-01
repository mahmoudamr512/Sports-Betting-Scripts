import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NBA"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/nba - data.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NBAScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(NBAScraper, self).__init__(executable_path="./geckodriver")
        self.playersStats = []
        self.gameLinks = []

    def openBetMGM(self) -> None:
        self.get("https://sports.pa.betmgm.com/en/sports/basketball-7/betting/usa-9/nba-6004")

    def getAllGamesLinks(self) -> None:
        events = WebDriverWait(self, wait_time).until(ec.presence_of_all_elements_located((By.TAG_NAME, "ms-six-pack-event")))

        for event in events:
            self.gameLinks.append(f"{event.find_element(By.TAG_NAME, 'a').get_attribute('href')}?market=3")

    def scrapGames(self, stats, startGame = 1, endGame = 'all') -> None:

        self.endGame = endGame
        self.startGame = startGame

        if endGame == 'all':
            self.endGame = len(self.gameLinks)

        counter = 1

        for link in self.gameLinks[(self.startGame)-1:self.endGame]:
            self.get(link)
            print(f"Scraping Game {counter}")
            self.scrapStats(stats)
            print("Done")
            print(self.playersStats)
            counter += 1

    def scrapStats(self, stats) -> None:
        mainDivs = WebDriverWait(self, wait_time).until(ec.presence_of_all_elements_located((By.CSS_SELECTOR, 'ms-option-panel.option-panel')))
        for i in range(len(mainDivs)):
            try:
                tabsBar = WebDriverWait(mainDivs[i], 10).until(ec.presence_of_element_located((By.CSS_SELECTOR, 'ul.tab-bar-container')))
                tabs = tabsBar.find_elements(By.XPATH, './li')
                for j in range(len(tabs)):
                    self.execute_script("arguments[0].scrollIntoView(true);", tabs[j])
                    if tabs[j].text.strip() in stats:
                        self.execute_script("arguments[0].click();", tabs[j])
                        playersDiv = mainDivs[i].find_element(By.CSS_SELECTOR,
                                                              '.option-group-container.player-props-container.triple')
                        self.execute_script("arguments[0].scrollIntoView(true);", playersDiv)
                        try:
                            showBtn = mainDivs[i].find_element(By.CSS_SELECTOR, "div.show-more-less-button")
                            if "expanded" not in showBtn.get_attribute('outerHTML'):
                                self.execute_script("arguments[0].click();", showBtn)
                                time.sleep(1)
                                playersDiv = mainDivs[i].find_element(By.CSS_SELECTOR,
                                                                      '.option-group-container.player-props-container.triple')
                        except Exception as ex:
                            pass

                        players = playersDiv.find_elements(By.CSS_SELECTOR, "div.attribute-key")
                        options = playersDiv.find_elements(By.TAG_NAME, 'ms-option')
                        for k in range(len(players)):
                            player = {
                                "Player Name": players[k].text,
                                "Stat": tabs[j].text.title(),
                                "Total": options[k*2].find_element(By.CLASS_NAME, "name").text,
                                "Over Odds": options[k*2].find_element(By.CLASS_NAME, "value").text,
                                "Under Odds": options[(k*2) + 1].find_element(By.CLASS_NAME, "value").text
                            }
                            self.playersStats.append(player)
                            print(player)
                        mainDivs = WebDriverWait(self, wait_time).until(
                            ec.presence_of_all_elements_located((By.TAG_NAME, 'ms-option-panel')))
                        tabsBar = WebDriverWait(mainDivs[i], 10).until(
                            ec.presence_of_element_located((By.CSS_SELECTOR, 'ul.tab-bar-container')))
                        tabs = tabsBar.find_elements(By.XPATH, './li')
            except:
                pass


with NBAScraper() as nba:
    print("Opening Bet MGM")
    nba.openBetMGM()
    print("Getting All Links")
    nba.getAllGamesLinks()
    print(f"Found {nba.gameLinks} games.")
    nba.scrapGames(["Points", "Rebounds", "Assists",
                    "Points+Assists+Rebounds", "Assists+Rebounds", "Points+Assists", "Points+Rebounds"],1,2)

    print("Extracting to Excel :) ")
    extractToExcel(nba.playersStats, "nba - data", "BetMGM Raw")
    print("Done, closing..")
