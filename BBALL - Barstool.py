import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

from selenium.webdriver.common.action_chains import ActionChains

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NBA"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/nba - data.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NBAScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        options = webdriver.FirefoxOptions()
        options.add_argument('--disable-blink-features=AutomationControlled')
        super(NBAScraper, self).__init__(executable_path="./geckodriver", options=options)
        self.playersStats = []
        self.actions = ActionChains(self)
        self.gameLinks = []

    def openBarstool(self) -> None:
        self.get("https://www.barstoolsportsbook.com/sports/basketball/nba")

    def getAllGamesLinks(self) -> None:
        """
            Get all game links
        """
        gameEvents = WebDriverWait(self, 30).until(ec.presence_of_all_elements_located((By.CSS_SELECTOR, '.container.wrap.event-row.match-row')))

        for game in gameEvents:
            try:
                self.gameLinks.append(
                    game.find_element(By.TAG_NAME, 'a').get_attribute('href')
                )
            except:
                pass

    def scrapTab(self, stats, tabDiv):
        time.sleep(3)

        innerDivs = tabDiv.find_elements(By.CLASS_NAME, "offer-item-wrapper")

        for div in innerDivs:
            self.execute_script("arguments[0].scrollIntoView(true);", div)

            divHeading = div.find_element(By.CSS_SELECTOR, ".ml-sp1.strongbody2").text.title()

            if divHeading in stats:
                rows = div.find_elements(By.CSS_SELECTOR, ".offer-row.flexbox.justify-center")

                for row in rows:
                    rowText = row.text.split("\n")

                    players = {
                        "Player Name": rowText[0],
                        "Stat": divHeading,
                        "Total": rowText[1].split(" ")[1],
                        "Over Odds": rowText[2],
                        "Under Odds": rowText[4]
                    }

                    print(players)

                    self.playersStats.append(players)


    def scrapGame(self, stats, link):
        self.get(
            link
        )

        time.sleep(5)

        tabs = WebDriverWait(self, 15).until(ec.presence_of_all_elements_located((By.CLASS_NAME, 'v-tab')))


        for i in range(len(tabs)):
            if tabs[i].text in ["All"]:
                self.execute_script("arguments[0].scrollIntoView(true);", tabs[i])
                tabs[i].click()
                self.execute_script("arguments[0].click()", tabs[i])
                time.sleep(3)
                self.scrapTab(stats, self.find_element(By.CSS_SELECTOR,'.v-window-item.active-tab').find_element(
                    By.XPATH, "./div/div[1]"
                ))

    def scrapGames(self, stats , startGame = 1, endGame= "all") -> None:
        self.startGame = startGame - 1
        self.endGame = endGame
        if endGame == "all":
            self.endGame = len(self.gameLinks)

        for i in range(self.startGame, self.endGame):
            print(f"Scraping Game {i+1}")
            self.scrapGame(stats, self.gameLinks[i])


with NBAScraper() as nba:
    print("Opening Barstool")
    nba.openBarstool()

    print("Getting All Games")
    nba.getAllGamesLinks()
    print(f"Found {len(nba.gameLinks)} games.")

    nba.scrapGames(["POINTS, REBOUNDS & ASSISTS BY THE PLAYER", "POINTS SCORED BY THE PLAYER", "ASSISTS BY THE PLAYER", "REBOUNDS BY THE PLAYER"],1,1)

    print("Extracting to Excel :) ")
    extractToExcel(nba.playersStats, "nba - data", "Barstool Raw")
    print("Done, closing..")
