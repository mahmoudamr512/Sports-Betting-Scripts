import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

from selenium.webdriver.common.action_chains import ActionChains

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NFL"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/{excelFile}.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NFLScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(NFLScraper, self).__init__(executable_path="./geckodriver")
        self.playersStats = []
        self.actions = ActionChains(self)
        self.gameLinks = []

    def openFaundel(self) -> None:
        self.get("https://pa.sportsbook.fanduel.com/football/nfl")

    def getAllGamesLinks(self) -> None:
        """
            Get all game links
        """
        gameEvents = WebDriverWait(self, 30).until(ec.presence_of_all_elements_located((By.XPATH, '//*[@id="root"]/div/div[2]/div[1]/div[2]/div[1]/div/div/div/div[2]/div/div')))

        for game in gameEvents:
            try:
                self.gameLinks.append(
                    game.find_element(By.TAG_NAME, 'a').get_attribute('href')
                )
            except:
                pass

    def scrapTab(self, stats, tabDiv):
        time.sleep(3)

        innerDivs = WebDriverWait(tabDiv, 30).until(ec.presence_of_all_elements_located((By.XPATH, "./div")))[1:]
        for i in range(len(innerDivs)):
            flag = False

            for stat in stats:
                if stat in innerDivs[i].text:
                    print(stat)
                    flag = True

            if flag:
                if i > 0:
                    time.sleep(0.25)
                    self.execute_script("arguments[0].click();", innerDivs[i].find_element(By.XPATH,'.//div[@role="button"]'))
                    time.sleep(3)

                try:
                    show_more = self.find_element(By.XPATH, './/span[contains(text(), "' + 'Show more' + '")]')
                    self.execute_script("arguments[0].scrollIntoView(true);", show_more)
                    time.sleep(0.5)
                    self.execute_script("arguments[0].click();", show_more)
                    time.sleep(2)
                except:
                    pass

                playersDiv = innerDivs[i].find_elements(By.XPATH, "./div/div/div[3]/div")

                for playerDiv in playersDiv:
                    overUnder = playerDiv.find_element(By.XPATH, "./div/div[2]")
                    player = {
                        "Player Name": playerDiv.find_element(By.XPATH, "./div/div[1]").text.strip(),
                        "Stat": innerDivs[i].find_element(By.XPATH, "./div/div/div/div").text,
                        "Total": overUnder.find_element(By.XPATH, './div[1]/span[1]').text.strip().split(" ")[1],
                        "Over Odds": overUnder.find_element(By.XPATH, './div[1]/span[2]').text.strip(),
                        "Under Odds": overUnder.find_element(By.XPATH, './div[2]/span[2]').text.strip()
                    }

                    self.playersStats.append(player)
                    print(player)
            self.execute_script("arguments[0].scrollIntoView(true);", innerDivs[0])


    def scrapGame(self, stats, link):
        self.get(
            link
        )

        tabs = WebDriverWait(self, 15).until(ec.presence_of_all_elements_located((By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div/div[2]/div[2]/div/div/div/div/div/nav/ul/li')))


        for i in range(len(tabs)):
            if tabs[i].text in ["Passing Props", "Receiving Props", "Rushing Props"]:
                self.execute_script("arguments[0].scrollIntoView(true);", tabs[i])
                tabs[i].click()
                print(tabs[i].text)
                time.sleep(2.5)
                self.scrapTab(stats, self.find_elements(By.XPATH,'//*[@style="flex-direction: column; overflow: hidden auto; display: flex; min-width: 0px;"]')[1])

    def scrapGames(self, stats , startGame = 1, endGame= "all") -> None:
        self.startGame = startGame - 1
        self.endGame = endGame
        if endGame == "all":
            self.endGame = len(self.gameLinks)

        for i in range(self.startGame, self.endGame):
            print(f"Scraping Game {i+1}")
            self.scrapGame(stats, self.gameLinks[i])


with NFLScraper() as nfl:
    print("Opening Faundel")
    nfl.openFaundel()

    print("Getting All Games")
    nfl.getAllGamesLinks()
    print(f"Found {len(nfl.gameLinks)} games.")

    nfl.scrapGames(["Player Passing Yds", "Player Receiving Yds", "Player Total Receptions", "Player Rushing Yds", "Player Rush Attempts"])

    print("Extracting to Excel :) ")
    extractToExcel(nfl.playersStats, "nfl - data", "Faundel")
    print("Done, closing..")
