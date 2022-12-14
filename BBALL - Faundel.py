import openpyxl
from selenium import webdriver
from openpyxl import load_workbook

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

from selenium.webdriver.common.action_chains import ActionChains

import pandas as pd
import sys
import os
import time

# Global Variables needed for the application
wait_time = 10
target_sport = "NBA"
master_list = []

def extractToExcel(list, excelFile, excelTab):
    path = f'/Users/ryanmccarroll/Google Drive/PyCharm Output/nba - data.xlsx'

    #path = f'{excelFile}.xlsx'

    writer = None

    if os.path.exists(f'{excelFile}.xlsx'):
        writer = pd.ExcelWriter(path, engine='openpyxl')
    else:
        openpyxl.Workbook().save(f"{excelFile}.xlsx")
        writer = pd.ExcelWriter(path, engine='openpyxl')

    book = load_workbook(path)
    writer.book = book

    try:
        book.remove(book[excelTab])
    except Exception as ex:
        pass

    result = pd.DataFrame(list)
    result = result[['Player Name', 'Stat', 'Total', 'Over Odds', 'Under Odds']]
    result.to_excel(writer, sheet_name=excelTab)
    writer.save()
    writer.close()
    book.save(path)
    book.close()

class NBAScraper(webdriver.Firefox):

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.quit()

    def __init__(self) -> None:
        super(NBAScraper, self).__init__(executable_path="./geckodriver")
        self.playersStats = []
        self.actions = ActionChains(self)
        self.gameLinks = []

    def openFaundel(self) -> None:
        self.get("https://pa.sportsbook.fanduel.com/navigation/nba")

    def getAllGamesLinks(self) -> None:
        """
            Get all game links
        """
        time.sleep(2)
        parent_divs =  WebDriverWait(self, 15)\
            .until(ec.visibility_of_all_elements_located(
            (By.XPATH,
             '//div[@style="flex-direction: column; overflow: hidden auto; display: flex; min-width: 0px;"]'
             )
        )
        )

        target_div = None
        for div in parent_divs:
            if "money" in div.text.lower():
                target_div = div

        gameEvents = target_div.find_elements(By.XPATH, './div')
        for game in gameEvents:
            try:
                self.gameLinks.append(
                    game.find_element(By.TAG_NAME, 'a').get_attribute('href')
                )
            except:
                pass

    def scrapTab(self, stats, tabDiv):

        innerDivs = WebDriverWait(tabDiv, 30).until(ec.presence_of_all_elements_located((By.XPATH, "./div")))[1:]
        for i in range(len(innerDivs)):
            flag = False

            for stat in stats:
                if stat in innerDivs[i].text:
                    print(stat)
                    flag = True

            if flag:
                if i > 0:
                    self.execute_script("arguments[0].click();", innerDivs[i].find_element(By.XPATH,'.//div[@role="button"]'))
                    time.sleep(1)

                try:
                    show_more = self.find_element(By.XPATH, './/span[contains(text(), "' + 'Show more' + '")]')
                    self.execute_script("arguments[0].scrollIntoView(true);", show_more)
                    time.sleep(0.5)
                    self.execute_script("arguments[0].click();", show_more)
                    time.sleep(0.5)
                except:
                    pass

                playersDiv = innerDivs[i].find_elements(By.XPATH, "./div/div/div[3]/div")

                for playerDiv in playersDiv:
                    overUnder = playerDiv.find_element(By.XPATH, "./div/div[2]")
                    player = {
                        "Player Name": playerDiv.find_element(By.XPATH, "./div/div[1]").text.strip(),
                        "Stat": innerDivs[i].find_element(By.XPATH, "./div/div/div/div").text,
                        "Total": overUnder.find_element(By.XPATH, './div[1]/span[1]').text.strip().split(" ")[1],
                        "Over Odds": overUnder.find_element(By.XPATH, './div[1]/span[2]').text.strip(),
                        "Under Odds": overUnder.find_element(By.XPATH, './div[2]/span[2]').text.strip()
                    }

                    self.playersStats.append(player)
                    print(player)
            self.execute_script("arguments[0].scrollIntoView(true);", innerDivs[0])


    def scrapGame(self, stats, link):
        self.get(
            link
        )

        tabs = WebDriverWait(self, 15).until(ec.presence_of_all_elements_located((By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div/div[2]/div[2]/div/div/div/div/div/nav/ul/li')))


        for i in range(len(tabs)):
            if tabs[i].text in ["Player Points", "Player Rebounds", "Player Assists", "Player Combos"]:
                self.execute_script("arguments[0].scrollIntoView(true);", tabs[i])
                tabs[i].click()
                print(tabs[i].text)
                time.sleep(0.5)
                self.scrapTab(stats, self.find_elements(By.XPATH,'//*[@style="flex-direction: column; overflow: hidden auto; display: flex; min-width: 0px;"]')[1])

    def scrapGames(self, stats , startGame = 1, endGame= "all") -> None:
        self.startGame = startGame - 1
        self.endGame = endGame
        if endGame == "all":
            self.endGame = len(self.gameLinks)

        for i in range(self.startGame, self.endGame):
            print(f"Scraping Game {i+1}")
            self.scrapGame(stats, self.gameLinks[i])


with NBAScraper() as nba:
    print("Opening Faundel")
    nba.openFaundel()

    print("Getting All Games")
    nba.getAllGamesLinks()
    print(f"Found {len(nba.gameLinks)} games.")

    nba.scrapGames(["Player Points", "Player Rebounds", "Player Assists", "Player Pts + Ast", "Player Pts + Reb", "Player Reb + Ast", "Player Pts + Reb + Ast"],1,1)

    print("Extracting to Excel :) ")
    extractToExcel(nba.playersStats, "nba - data", "Fanduel Raw")
    print("Done, closing..")
